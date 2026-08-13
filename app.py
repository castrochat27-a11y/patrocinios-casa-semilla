import io
import os
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file
from sqlalchemy import (
    Column,
    DateTime,
    Integer,
    Numeric,
    String,
    Text,
    create_engine,
)
from sqlalchemy.orm import declarative_base, scoped_session, sessionmaker

import catalogos

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Conexión a la base de datos
# Usa Postgres (Supabase/Render) si existe DATABASE_URL; si no, SQLite local.
# ---------------------------------------------------------------------------
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if not DATABASE_URL:
    DATABASE_URL = "sqlite:///" + os.path.join(BASE_DIR, "patrocinios.db")

engine_kwargs = {"pool_pre_ping": True}
if DATABASE_URL.startswith("sqlite"):
    engine_kwargs["connect_args"] = {"check_same_thread": False}

engine = create_engine(DATABASE_URL, **engine_kwargs)
SessionLocal = scoped_session(sessionmaker(bind=engine, autoflush=False))
Base = declarative_base()

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Modelos (una tabla por sección del Excel)
# ---------------------------------------------------------------------------
class PosiblePatrocinio(Base):
    __tablename__ = "posibles_patrocinios"

    id = Column(Integer, primary_key=True)
    empresa = Column(String(300), nullable=False)
    contacto = Column(Text, default="")
    clasificacion = Column(String(100), default="")
    encargado = Column(String(150), default="")
    lista = Column(String(100), default="")
    valor_aproximado = Column(Numeric(14, 2))
    observaciones = Column(Text, default="")
    creado_en = Column(DateTime, default=datetime.utcnow)
    actualizado_en = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    CAMPOS = [
        "empresa",
        "contacto",
        "clasificacion",
        "encargado",
        "lista",
        "valor_aproximado",
        "observaciones",
    ]
    ENCABEZADOS = [
        "#",
        "Empresa",
        "Contacto",
        "Clasificación",
        "Encargado",
        "Lista",
        "Valor aproximado",
        "Observaciones",
        "Registrado",
    ]


class EstadoPatrocinio(Base):
    __tablename__ = "estado_patrocinios"

    id = Column(Integer, primary_key=True)
    empresa = Column(String(300), nullable=False)
    contacto = Column(Text, default="")
    estado = Column(String(100), default="")
    encargado = Column(String(150), default="")
    descripcion = Column(Text, default="")
    tipo_patrocinio = Column(String(200), default="")
    valor_aproximado = Column(Numeric(14, 2))
    creado_en = Column(DateTime, default=datetime.utcnow)
    actualizado_en = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    CAMPOS = [
        "empresa",
        "contacto",
        "estado",
        "encargado",
        "descripcion",
        "tipo_patrocinio",
        "valor_aproximado",
    ]
    ENCABEZADOS = [
        "#",
        "Empresa",
        "Contacto",
        "Estado",
        "Encargado",
        "Descripción de la posible donación",
        "Tipo de Patrocinio",
        "Valor aproximado",
        "Registrado",
    ]


class RegistroDonacion(Base):
    __tablename__ = "registro_donaciones"

    id = Column(Integer, primary_key=True)
    empresa = Column(String(300), nullable=False)
    contacto = Column(Text, default="")
    encargado = Column(String(150), default="")
    tipo_donacion = Column(String(200), default="")
    descripcion = Column(Text, default="")
    valor_aproximado = Column(Numeric(14, 2))
    asignacion = Column(String(100), default="")
    tipo_patrocinio = Column(String(200), default="")
    observaciones = Column(Text, default="")
    creado_en = Column(DateTime, default=datetime.utcnow)
    actualizado_en = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    CAMPOS = [
        "empresa",
        "contacto",
        "encargado",
        "tipo_donacion",
        "descripcion",
        "valor_aproximado",
        "asignacion",
        "tipo_patrocinio",
        "observaciones",
    ]
    ENCABEZADOS = [
        "#",
        "Empresa",
        "Contacto",
        "Encargado",
        "Tipo de Donación",
        "Descripción de la donación",
        "Valor aproximado",
        "Asignación",
        "Tipo de Patrocinio",
        "Observaciones adicionales",
        "Registrado",
    ]


SECCIONES = {
    "posibles": {
        "modelo": PosiblePatrocinio,
        "titulo": "Posibles Patrocinios",
        "hoja": "Posibles Patrocinios",
    },
    "estado": {
        "modelo": EstadoPatrocinio,
        "titulo": "Estado de Patrocinios",
        "hoja": "Estado Patrocinios",
    },
    "donaciones": {
        "modelo": RegistroDonacion,
        "titulo": "Registro de Donaciones",
        "hoja": "Registro Donaciones",
    },
}

Base.metadata.create_all(engine)


@app.teardown_appcontext
def cerrar_sesion(exception=None):
    SessionLocal.remove()


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def limpiar_valor(valor):
    """Convierte el valor aproximado a número; devuelve None si viene vacío."""
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None
    texto = (
        texto.replace("₡", "")
        .replace("$", "")
        .replace(" ", "")
        .replace(",", "")
    )
    try:
        numero = float(texto)
    except ValueError:
        return None
    if numero < 0:
        return None
    return numero


def a_dict(registro, modelo):
    datos = {"id": registro.id}
    for campo in modelo.CAMPOS:
        valor = getattr(registro, campo)
        if campo == "valor_aproximado":
            datos[campo] = float(valor) if valor is not None else None
        else:
            datos[campo] = valor or ""
    datos["creado_en"] = (
        registro.creado_en.strftime("%d/%m/%Y") if registro.creado_en else ""
    )
    return datos


def aplicar_datos(registro, modelo, datos, parcial=False):
    for campo in modelo.CAMPOS:
        if parcial and campo not in datos:
            continue
        valor = datos.get(campo, "")
        if campo == "valor_aproximado":
            setattr(registro, campo, limpiar_valor(valor))
        else:
            setattr(registro, campo, (str(valor).strip() if valor is not None else ""))
    return registro


def obtener_seccion(clave):
    seccion = SECCIONES.get(clave)
    if seccion is None:
        return None
    return seccion


# ---------------------------------------------------------------------------
# Vistas
# ---------------------------------------------------------------------------
@app.route("/")
def inicio():
    return render_template("index.html", cat=catalogos)


@app.route("/api/<seccion>", methods=["GET"])
def listar(seccion):
    info = obtener_seccion(seccion)
    if info is None:
        return jsonify({"error": "Sección no encontrada."}), 404

    modelo = info["modelo"]
    db = SessionLocal()
    consulta = db.query(modelo)

    buscar = request.args.get("buscar", "").strip()
    encargado = request.args.get("encargado", "").strip()
    filtro = request.args.get("filtro", "").strip()

    if buscar:
        consulta = consulta.filter(modelo.empresa.ilike(f"%{buscar}%"))
    if encargado:
        consulta = consulta.filter(modelo.encargado == encargado)
    if filtro:
        if seccion == "posibles":
            consulta = consulta.filter(modelo.clasificacion == filtro)
        elif seccion == "estado":
            consulta = consulta.filter(modelo.estado == filtro)
        elif seccion == "donaciones":
            consulta = consulta.filter(modelo.asignacion == filtro)

    registros = consulta.order_by(modelo.id.desc()).all()
    return jsonify([a_dict(r, modelo) for r in registros])


@app.route("/api/<seccion>", methods=["POST"])
def crear(seccion):
    info = obtener_seccion(seccion)
    if info is None:
        return jsonify({"error": "Sección no encontrada."}), 404

    datos = request.get_json(silent=True) or {}
    if not str(datos.get("empresa", "")).strip():
        return jsonify({"error": "El nombre de la empresa es obligatorio."}), 400

    modelo = info["modelo"]
    db = SessionLocal()
    registro = aplicar_datos(modelo(), modelo, datos)
    db.add(registro)
    db.commit()
    db.refresh(registro)
    return jsonify(a_dict(registro, modelo)), 201


@app.route("/api/<seccion>/<int:registro_id>", methods=["PUT"])
def actualizar(seccion, registro_id):
    info = obtener_seccion(seccion)
    if info is None:
        return jsonify({"error": "Sección no encontrada."}), 404

    modelo = info["modelo"]
    db = SessionLocal()
    registro = db.get(modelo, registro_id)
    if registro is None:
        return jsonify({"error": "Registro no encontrado."}), 404

    datos = request.get_json(silent=True) or {}
    if "empresa" in datos and not str(datos.get("empresa", "")).strip():
        return jsonify({"error": "El nombre de la empresa es obligatorio."}), 400

    aplicar_datos(registro, modelo, datos, parcial=True)
    db.commit()
    db.refresh(registro)
    return jsonify(a_dict(registro, modelo))


@app.route("/api/<seccion>/<int:registro_id>", methods=["DELETE"])
def borrar(seccion, registro_id):
    info = obtener_seccion(seccion)
    if info is None:
        return jsonify({"error": "Sección no encontrada."}), 404

    modelo = info["modelo"]
    db = SessionLocal()
    registro = db.get(modelo, registro_id)
    if registro is None:
        return jsonify({"error": "Registro no encontrado."}), 404

    db.delete(registro)
    db.commit()
    return "", 204


@app.route("/api/resumen")
def resumen():
    db = SessionLocal()
    total_donado = 0.0
    for registro in db.query(RegistroDonacion).all():
        if registro.valor_aproximado is not None:
            total_donado += float(registro.valor_aproximado)
    return jsonify(
        {
            "posibles": db.query(PosiblePatrocinio).count(),
            "estado": db.query(EstadoPatrocinio).count(),
            "donaciones": db.query(RegistroDonacion).count(),
            "total_donado": total_donado,
        }
    )


@app.route("/exportar")
def exportar():
    """Genera el archivo Excel con las tres hojas y todos los registros."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    db = SessionLocal()
    wb = Workbook()
    wb.remove(wb.active)

    encabezado_fill = PatternFill("solid", fgColor="1F4E79")
    encabezado_font = Font(color="FFFFFF", bold=True)

    for clave, info in SECCIONES.items():
        modelo = info["modelo"]
        ws = wb.create_sheet(title=info["hoja"])
        ws.append(modelo.ENCABEZADOS)

        for celda in ws[1]:
            celda.fill = encabezado_fill
            celda.font = encabezado_font
            celda.alignment = Alignment(horizontal="center", vertical="center")

        registros = db.query(modelo).order_by(modelo.id).all()
        for indice, registro in enumerate(registros, start=1):
            fila = [indice]
            for campo in modelo.CAMPOS:
                valor = getattr(registro, campo)
                if campo == "valor_aproximado":
                    fila.append(float(valor) if valor is not None else None)
                else:
                    fila.append(valor or "")
            fila.append(
                registro.creado_en.strftime("%d/%m/%Y") if registro.creado_en else ""
            )
            ws.append(fila)

        # Formato de moneda en la columna "Valor aproximado"
        if "Valor aproximado" in modelo.ENCABEZADOS:
            columna = modelo.ENCABEZADOS.index("Valor aproximado") + 1
            for fila_num in range(2, ws.max_row + 1):
                ws.cell(row=fila_num, column=columna).number_format = '"₡"#,##0'

        anchos = [6, 34, 30, 22, 22, 34, 18, 30, 14, 24, 14]
        for i, encabezado in enumerate(modelo.ENCABEZADOS, start=1):
            ancho = anchos[i - 1] if i - 1 < len(anchos) else 20
            ws.column_dimensions[get_column_letter(i)].width = ancho

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(modelo.ENCABEZADOS))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"Registro_Patrocinios_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    puerto = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=puerto, debug=True)
